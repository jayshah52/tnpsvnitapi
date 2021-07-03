import datetime
import json
from io import BytesIO

import openpyxl as openpyxl
from django.core.mail import send_mail
from django.http import HttpResponse
from django.shortcuts import render
import pandas as pd
# Create your views here.
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from rest_framework import viewsets, filters, pagination
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.reverse import reverse
from .serializers import JobSerializer, ShortlistSerializer
from .models import Job, Shortlist
from student.models import Student

from student.serializers import StudentSerializer

from tnp import settings


class JobPagination(PageNumberPagination):
    page_size = 25

class JobStudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated, ]
    pagination_class = JobPagination
    filter_backends = (filters.SearchFilter,)
    search_fields = ('department', 'name', 'roll_no')

    def get_queryset(self):

        id = self.request.GET.get('id')
        job = Job.objects.get(id=id)
        print(job.users.all())
        users = job.users.all()
        filtered_queryset = self.filter_queryset(users)
        return filtered_queryset

    # @action(detail=False, methods=['get'])
    # def get_all_users(self, request, *args, **kwargs):
    #     id = request.query_params.get('id')
    #     job = Job.objects.get(id=id)
    #     print(job.users.all())
    #     users = job.users.all()
    #     # filtered_queryset = self.filter_queryset(users)
    #     # return filtered_queryset
    #     page = self.paginate_queryset(users)
    #     if page is not None:
    #         serializer = StudentSerializer(page, many=True)
    #         return self.get_paginated_response(serializer.data)
    #     serializer = StudentSerializer(users, many=True)
    #     return Response(serializer.data)

class ShortlistViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated, ]
    pagination_class = JobPagination
    filter_backends = (filters.SearchFilter,)
    search_fields = ('name', 'roll_no','department','graduation_year')

    def get_queryset(self):
        id = self.request.GET.get('id')
        number = self.request.GET.get('number')
        job = Job.objects.get(id=id)
        # shortlist = Shortlist.objects.filter(job = job, number = number).first()
        shortlist, created = Shortlist.objects.get_or_create(job=job, number=number)
        users = shortlist.users.all()
        # print(shortlist)
        # print(users)
        # for user in users:
        #     print(user.name)
        # print("SHORTLIST", shortlist)
        filtered_queryset = self.filter_queryset(users)
        return filtered_queryset

    @action(detail=False, methods=['post'])
    def add_student(self, request, *args, **kwargs):
        id = request.data.get('id')
        number = request.data.get('number')
        username = request.data.get('username')
        job = Job.objects.get(id=id)
        print("HELLOOOOOO",id, number, username)
        shortlist, created = Shortlist.objects.get_or_create(job = job, number=number)
        student = Student.objects.get(username = username)
        shortlist.users.add(student)
        return Response("Added student successfully!")

    @action(detail=False, methods=['post'])
    def remove_student(self, request, *args, **kwargs):
        id = request.data.get('id')
        number = request.data.get('number')
        username = request.data.get('username')
        job = Job.objects.get(id=id)
        # shortlist = Shortlist.objects.filter(job=job, number=number).first()
        student = Student.objects.get(username=username)
        for num in range(number,5):
            shortlist = Shortlist.objects.filter(job=job, number=num).first()
            shortlist.users.remove(student)
        return Response("Removed Successfully!")

class JobViewSet(viewsets.ModelViewSet):
    queryset = Job.objects.all()
    serializer_class = JobSerializer
    permission_classes = [IsAuthenticated,]
    pagination_class = JobPagination
    filter_backends = (filters.SearchFilter,)
    search_fields = ('for_departments', 'name', 'role')
    def get_queryset(self):
        return self.queryset

    @action(detail=False, methods=['post'])
    def create_job(self, request, *args, **kwargs):
        data = json.loads(request.data.get('data'))
        print(data)
        print(request.data.get('jd'))
        jd = request.data.get('jd')
        # print("ada", request.data.get('data'))
        # student = Student.objects.get(username = request.user)
        # data['created_by'] = request.user.username
        Job.objects.create(jd=jd,**data)
        # subject = 'New Job Posted'
        # message = f'Hi, A new Job has been posted'
        # email_from = 'jaydragon70@gmail.com'
        # recipient_list = ['vidt007@gmail.com', ]
        # send_mail(subject, message, email_from, recipient_list)
        return Response("Job Created successfully")

    @action(detail=False, methods=['get'])
    def get_job_info(self, request, *args, **kwargs):
        id = request.query_params.get('id')
        job = Job.objects.get(id=id)
        serializer = JobSerializer(job)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def update_job(self, request, *args, **kwargs):
        data = request.data
        id = request.query_params.get('id')
        print(id)
        print("data", request.data)
        # print("sad", **request.data)
        print(Job.objects.filter(id=id))
        Job.objects.filter(id=id).update(**request.data)
        job = Job.objects.get(id=id)
        print("NJOB",job)
        serializer = JobSerializer(job, context={'request': request})
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def get_all_jobs(self, request, *args, **kwargs):
        student = Student.objects.get(username = request.user)
        dept = student.department
        year = student.graduation_year
        now = datetime.datetime.now()
        jobs = self.queryset.filter(apply_by__gte=now,grad_year = year, for_departments__contains = dept)
        # print(jobs)
        serializer = self.serializer_class(jobs, many=True, context={'request': request})
        # print("WORKS", serializer.data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def job_detail(self, request, *args, **kwargs):
        id = request.query_params.get('id')
        job = Job.objects.get(id = id)
        serializer = self.serializer_class(job)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def delete(self, request, *args, **kwargs):
        id = request.data.get('id')
        Job.objects.filter(id=id).delete()
        return Response("Deleted Succesfully!")

    @action(detail=False, methods=['post'])
    def register(self, request, *args, **kwargs):
        student = Student.objects.get(username = request.user)
        id = request.data.get('id')
        job = Job.objects.get(id=id)
        job.users.add(student)
        serializer = self.serializer_class(job)
        return Response(serializer.data)



    @action(detail=False, methods=['get'])
    def get_my_jobs(self, request, *args, **kwargs):
        student = Student.objects.get(username = request.user)
        jobs = student.job_set.all()
        serializer = self.serializer_class(jobs, many=True, context={'request': request})
        # print("WORKS", serializer.data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def download_excel(self,request,*args,**kwargs):
        wb = openpyxl.Workbook()
        sheet = wb.active
        ws = wb.create_sheet('Sorting')
        id = request.query_params.get('id')
        dept = request.query_params.get('dept')
        job = Job.objects.get(id=id)
        users = job.users.all().filter(department = dept)
        # print(users)
        def style_cell(cell, val, size=12):
            sheet[cell] = val
            sheet[cell].font = Font(size=size, bold=True, name='Arial')
            sheet[cell].alignment = Alignment(horizontal='center', vertical='center')
        def write_cell(cell,val):
            ws[cell] = val
        sheet.merge_cells('A1:Q1')
        # ws['A1'].value = "HELLO WORLD"
        sheet['A1'].value = 'SARDAR VALLABHBHAI NATIONAL INSTITUTE OF TECHNOLOGY (SVNIT), SURAT (GUJARAT)'
        sheet['A1'].font = Font(size=20, bold=True, name='Arial')
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A2:Q2')
        style_cell('A2', 'TRAINING & PLACEMENT SECTION (T&P)', 20)

        sheet.merge_cells('A3:Q3')
        style_cell('A3', 'UG (B.Tech.)  :   Engineering  :  2021-22 Batch', 16)
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 25
        sheet.row_dimensions[3].height = 20
        col_names = ['Sr No.', 'Adm. / Roll No.', 'Name', 'Gender', 'DOB', 'Home Town', 'Current Location']
        i = 0
        for al in 'ABCDEFG':
            sheet.merge_cells('{}4{}{}5'.format(al, ':', al))
            style_cell('{}4'.format(al), col_names[i])
            ws['{}1'.format(al)] = col_names[i]
            i += 1

        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 18

        sheet.merge_cells('H4:I4')
        sheet.merge_cells('J4:O4')
        ws['H1'] = '10th'
        ws['I1'] = '12th'
        style_cell('H4', '%')

        style_cell('J4', 'CGPA after Semester')

        style_cell('H5', 'X')
        style_cell('I5', 'XII')

        col_names = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VI']
        i = 0
        for al in 'JKLMNO':
            style_cell('{}5'.format(al), col_names[i])
            ws['{}1'.format(al)] = col_names[i]
            i += 1
        ws['P1'].value = 'Email ID'
        ws['Q1'].value = 'Mobile No.'
        sheet.merge_cells('P4:P5')
        sheet.merge_cells('Q4:Q5')

        style_cell('P4', 'E-mail ID')
        style_cell('Q4', 'Mobile No.')

        sheet.row_dimensions[4].height = 25
        sheet.column_dimensions['P'].width = 15
        sheet.column_dimensions['Q'].width = 18

        i = 0
        for row in range(2,2+len(users)):
            write_cell('A{}'.format(row), i + 1)
            style_cell('A{}'.format(row+4), i + 1)
            write_cell('B{}'.format(row), users[i].roll_no)
            write_cell('C{}'.format(row), users[i].name)
            write_cell('D{}'.format(row), users[i].gender)
            write_cell('E{}'.format(row), users[i].dob)
            write_cell('F{}'.format(row), users[i].hometown)
            write_cell('G{}'.format(row), users[i].current_loc)
            write_cell('H{}'.format(row), users[i].percentage_ten)
            write_cell('I{}'.format(row), users[i].percentage_twelve)
            write_cell('J{}'.format(row), users[i].cgpa_s1)
            write_cell('K{}'.format(row), users[i].cgpa_s2)
            write_cell('L{}'.format(row), users[i].cgpa_s3)
            write_cell('M{}'.format(row), users[i].cgpa_s4)
            write_cell('N{}'.format(row), users[i].cgpa_s5)
            write_cell('O{}'.format(row), users[i].cgpa_s6)
            write_cell('P{}'.format(row), users[i].personal_mail)
            write_cell('Q{}'.format(row), users[i].phone_no)
            i += 1

        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data=data, columns=columns)
        df.sort_values(['VI'], ascending=False, inplace=True, ignore_index=True)
        # print(df)
        i = 0
        for row in range(6, 6 + len(users)):
            print(df['Adm. / Roll No.'][i])
            style_cell('B{}'.format(row), df['Adm. / Roll No.'][i])
            style_cell('C{}'.format(row), df['Name'][i])
            style_cell('D{}'.format(row), df['Gender'][i])
            style_cell('E{}'.format(row), df['DOB'][i])
            style_cell('F{}'.format(row), df['Home Town'][i])
            style_cell('G{}'.format(row), df['Current Location'][i])
            style_cell('H{}'.format(row), df['10th'][i])
            style_cell('I{}'.format(row), df['12th'][i])
            style_cell('J{}'.format(row), df['I'][i])
            style_cell('K{}'.format(row), df['II'][i])
            style_cell('L{}'.format(row), df['III'][i])
            style_cell('M{}'.format(row), df['IV'][i])
            style_cell('N{}'.format(row), df['V'][i])
            style_cell('O{}'.format(row), df['VI'][i])
            style_cell('P{}'.format(row), df['Email ID'][i])
            style_cell('Q{}'.format(row), df['Mobile No.'][i])
            i += 1
        # print("TRy", sdf.head())
        wb.remove(wb.get_sheet_by_name('Sorting'))
        wb.save('tnpsvnit.xlsx')
        wb.close()
        output = BytesIO()
        wb.save(output)
        # print("Hello",bio.seek(0))
        xlsx_data = output.getvalue()
        # print(xlsx_data.get_value())
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = "attachment; filename=test.xlsx"
        # print("RC",response['Content-Disposition'])
        response.write(xlsx_data)
        return response



